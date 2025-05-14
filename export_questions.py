import json
import openpyxl

wb = openpyxl.load_workbook('vmcev12q gemini.xlsx')
ws = wb.active
questions = []
headers = [cell.value for cell in ws[1]]
for row in ws.iter_rows(min_row=2, values_only=True):
    q = dict(zip(headers, row))
    # Build the question object for the quiz
    question_obj = {
        "stage": f"Question {len(questions)+1}",
        "question": q.get("Question", "").strip(),
        "options": [
            f"A. {q.get('A', '').strip()}",
            f"B. {q.get('B', '').strip()}",
            f"C. {q.get('C', '').strip()}",
            f"D. {q.get('D', '').strip()}"
        ],
        "correct": [q.get("Correct", "A").strip().upper()],
        "explanation": q.get("Explanation", "").strip()
    }
    questions.append(question_obj)

with open('vmce_questions.json', 'w', encoding='utf-8') as f:
    json.dump(questions, f, ensure_ascii=False, indent=2)
print(f"Exported {len(questions)} questions to vmce_questions.json")
