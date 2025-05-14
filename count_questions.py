import openpyxl

wb = openpyxl.load_workbook('vmcev12q gemini.xlsx')
ws = wb.active
questions = []
headers = [cell.value for cell in ws[1]]
for row in ws.iter_rows(min_row=2, values_only=True):
    q = dict(zip(headers, row))
    questions.append(q)
print(len(questions))
